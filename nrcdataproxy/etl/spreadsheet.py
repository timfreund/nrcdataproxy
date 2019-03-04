from datetime import datetime, time
from decimal import Decimal
from optparse import OptionParser
import agate
import agateexcel
import agatesql
import mimetypes
import os
import openpyxl
import sys

class SpreadsheetExtractor():
    mimetype = None

    location_keys = ['lat_', 'long_',]
    mapped_names = {'material_inv0lved_cr': 'material_involved_cr'}
    negatives = ['N', 'NO']
    positives = ['Y', 'YES']
    multi_entry_sheets = ['MATERIAL_INVOLVED', 
                          'MATERIAL_INV0LVED_CR', 
                          'TRAINS_DETAIL',
                          'DERAILED_UNITS',
    ]
    
    def __init__(self, filename):
        self.filename = filename

    def mapped_name(self, name):
        return self.mapped_names.get(name, name)

    def seperate_location_data(self, record):
        location_data = {}
        for k, v in record.items():
            if k.startswith("location_"):
                location_data[k.replace("location_", "")] = v
                del record[k]
            for lk in self.location_keys:
                if k.startswith(lk):
                    location_data[k] = v
                    del record[k]
        record['location'] = location_data
        return record
        
    def scrub_data(self, record):
        newrec = {}
        for k, v in record.items():
            newrec[k.lower()] = v
        record = newrec
        for k in list(record.keys()):
            v = record[k]
            if isinstance(v, dict):
                if len(v.keys()):
                    record[k] = self.scrub_data(v)
                else:
                    del record[k]
            if isinstance(v, list):
                if not len(v):
                    del record[k]
                else:
                    newlist = []
                    for vprime in v:
                        newlist.append(self.scrub_data(vprime))
                    record[k] = newlist
            if isinstance(v, datetime):
                record[k] = v.isoformat()
            if isinstance(v, Decimal):
                try:
                    record[k] = int(v)
                except ValueError:
                    record[k] = float(v)
            if isinstance(v, str) and v == "" or v is None:
                del record[k]
            if v in self.negatives:
                record[k] = False
            if v in self.positives:
                record[k] = True
        return record

    def extract_data(self, repository):
        for record in self.records:
            repository.save(self.seperate_location_data(self.scrub_data(record)))

class XlsxAgateExtractor(SpreadsheetExtractor):
    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, *args):
        SpreadsheetExtractor.__init__(self, *args)
        self.load_metadata()
        # import IPython; IPython.embed()

    def load_metadata(self):
        self.workbook = openpyxl.load_workbook(self.filename,
                                               read_only=True, data_only=True)
        self.sheetnames = self.workbook.sheetnames
        self.sheets = {}

        for sheetname in self.sheetnames:
            self.sheets[sheetname] = agate.Table.from_xlsx(self.filename,
                                                           sheet=sheetname,
                                                           row_names=lambda r: '%(SEQNOS)s' % (r))

    @property
    def records(self):
        calls = self.sheets['CALLS']
        for seqnos in calls.columns[0].values():
            if seqnos is not None:
                yield self.get_record(seqnos)

    def get_record(self, seqnos):
        # TODO ponder seqnos types
        # we have named rows, but that involves casting the seqnos to a string
        # We can't used the named rows for sheets with multiple entries per SEQNOS
        # so we end up doing a sheet.where to get those multi-entry sheets.  To do
        # that we cast back to an int.  It's kind ugly.
        #
        # maybe we use sheet.where for all cases, but only return in a list for confirmed
        # multi-entry sheets?
        seqnos = '%d' % seqnos
        call_record = self.sheets['CALLS'].rows[seqnos].dict()
        for sn in self.sheetnames:
            if sn != 'CALLS':
                sheet = self.sheets[sn]
                if sn in self.multi_entry_sheets:
                    sub_record_list = []
                    subset = sheet.where(lambda r: r['SEQNOS'] == int(seqnos))
                    for r in subset.rows:
                        sub_record_list.append(r.dict())
                    call_record[sn] = sub_record_list
                else:
                    try:
                        sub_record = self.sheets[sn].rows[seqnos]
                        call_record[sn] = sub_record.dict()
                    except KeyError:
                        pass
        return call_record

class XlsxExtractor(SpreadsheetExtractor):
    mimetype = 'disable-application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, *args):
        SpreadsheetExtractor.__init__(self, *args)
        self.workbook = openpyxl.reader.excel.load_workbook(self.filename)
        self.metadata = self.load_metadata()

    def load_metadata(self):
        metadata = {'layout': [],
                    'positions': {}}
        metadata['current_row'] = 1

        for sheet in self.workbook.worksheets:
            name = sheet.title
            columns = []
            for x in range(0, len(sheet.column_dimensions.keys())):
                # CY11 has a null value in the last column of INCIDENT_DETAILS.
                # We won't take any column that has a null header
                if sheet.cell(row=0, column=x).value is not None:
                    columns.append(sheet.cell(row=0, column=x).value)

            if len(columns):
                if columns[0] == u' ':
                    columns[0] = u'SEQNOS'

                sheet_keys = {}
                for row, cell in enumerate(sheet.columns[0]):
                    rows = sheet_keys.get(cell.value, [])
                    rows.append(row)
                    sheet_keys[cell.value] = rows
                    
                metadata['layout'].append((name, columns))
                metadata['positions'][name] = sheet_keys

        return metadata

    @property
    def records(self):
        data = {}

        name, columns = self.metadata['layout'][0]
        sheet = self.workbook.get_sheet_by_name(name)
        rowx = self.metadata['current_row']

        if rowx == -1:
            raise StopIteration

        for colx, col_name in enumerate(columns):
            cell = sheet.cell(row=rowx, column=colx)
            value = cell.value
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, time):
                value = None
            elif isinstance(value, unicode):
                value = value.strip()

            data[col_name.lower()] = value

        if sheet.cell(row=rowx+1, column=0).value != None:
            self.metadata['current_row'] = rowx + 1
        else:
            self.metadata['current_row'] = -1

        for name, columns in self.metadata['layout'][1:]:
            detail_data = self.incident_details(name, columns, data['seqnos'])

            if len(detail_data) > 1 and name not in self.multi_entry_sheets:
                msg = "%s: multiple entries in a single entry sheet: %s" % (data['seqnos'],
                                                                            name)
                raise Exception(msg)

            for dd in detail_data:
                if dd.has_key('SEQNOS'):
                    del dd['SEQNOS']

            if name == 'INCIDENT_COMMONS' and len(detail_data):
                for k, v in detail_data[0].items():
                    data[k.lower()] = v
            else:
                lname = self.mapped_name(name.lower())
                if name in self.multi_entry_sheets:
                    data[lname] = []
                    for dd in detail_data:
                        data[lname].append(dd)
                elif len(detail_data):
                    data[lname] = detail_data[0]

        yield data

    def incident_details(self, sheet_name, columns, seqnos):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        all_data = []

        if self.metadata['positions'][sheet_name].has_key(seqnos):
            for sheet_row in self.metadata['positions'][sheet_name][seqnos]:
                data = {}
                for coly, col_name in enumerate(columns):
                    cell = sheet.cell(row=sheet_row, column=coly)
                    value = cell.value
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, time):
                        value = None
                    elif isinstance(value, unicode):
                        value = value.strip()
                
                    data[col_name.lower()] = value
                del data['seqnos']
                all_data.append(data)
        return all_data
        
extractors = [
    XlsxAgateExtractor
    ]

def env_db_string():
    user = os.environ.get('POSTGRES_USER', 'user_not_set')
    password = os.environ.get('POSTGRES_PASSWORD', 'password_not_set')
    database = os.environ.get('POSTGRES_DB', 'database_not_set')
    host = os.environ.get('POSTGRES_HOST', 'db')
    port = os.environ.get('POSTGRES_port', '5432')

    return "postgresql://%s:%s@%s:%s/%s" % (user, password, host,
                                            port, database)

def extract_xlsx_to_sql(filename, sqlurl):
    workbook = openpyxl.load_workbook(filename,
                                      read_only=True, data_only=True)
    sheetnames = workbook.sheetnames

    # Agate makes an educated guess at data types, but sometimes it
    # guesses wrong.  This is especially possible when dealing with
    # multiple data files like we are in this project.  We're able
    # to override the type inference system here
    #
    # Many columns are sparsely populated: mostly blank, with maybe
    # a 1 or a 0 once in a while.  Agate will assume those are boolean
    # columns, but once in a while there'll be a number.
    specified_types = {}
    specified_types['CALLS'] = {
        'RESPONSIBLE_ZIP': agate.Text(), # CY90 '-    '
    }
    specified_types['DERAILED_UNITS'] = {
        'POSITION_IN_TRAIN': agate.Text(), # U
    }
    specified_types['INCIDENTS'] = {
        'AIRCRAFT_HANGER': agate.Text(),
        'BERTH_SLIP_NUMBER': agate.Text(),
        'BRAKEMAN_TESTING': agate.Number(),
        'DATE_TIME_NORMAL_SERVICE': agate.DateTime(),
        'OTHER_EMPLOYEE_TESTING': agate.Number(),
        'RCL_OPERATOR_TESTING': agate.Number(),
        'SIGNALMAN_TESTING': agate.Number(),
        'TRAIN_DISPATCHER_TESTING': agate.Number(),
        'TRAINMAN_TESTING': agate.Number(),
        'UNKNOWN_TESTING': agate.Text(), # CY10, YES
        'YARD_FOREMAN_TESTING': agate.Number(),
    }
    specified_types['INCIDENT_COMMONS'] = {
        'INCIDENT_LOCATION': agate.Text(),
        'LAT_QUAD': agate.Text(),
        'LONG_QUAD': agate.Text(),
        'LOCATION_ZIP': agate.Text(),
    }
    specified_types['INCIDENT_DETAILS'] = {
        'AIR_CLOSURE_TIME': agate.Number(),
        'AIR_CORRIDOR_CLOSED': agate.Text(),
        'AIR_CORRIDOR_DESC': agate.Text(),
        'ANY_FATALITIES': agate.Text(), # CY96, U
        'ANY_INJURIES': agate.Text(), # CY96, U
        'COMMUNITY_IMPACT': agate.Text(),
        'EMPL_FATALITY': agate.Number(),
        'ESTIMATED_DURATION_OF_RELEASE': agate.Text(), # CY14, on-going
        'FIRE_EXTINGUISHED': agate.Text(), # CY96, U
        'OFFSHORE': agate.Text(), # CY08, U (unknown?)
        'PASS_FATALITY': agate.Number(),
        'RELEASE_RATE': agate.Text(), # CY15, UNK
        'RELEASE_SECURED': agate.Text(), # CY15, UNK
        'ROAD_CLOSED': agate.Text(), # CY15, UNK
        'TRACK_CLOSED': agate.Text(), # CY96, U
        'TRACK_CLOSURE_TIME': agate.Number(), # CY99, 1
        'WATERWAY_CLOSED': agate.Text(), # CY96, U
        'WATERWAY_CLOSURE_TIME': agate.Text(), # CY96, U
        'WATERWAY_CLOSURE_TIME': agate.Number(), # CY12
    }
    specified_types['MATERIAL_INVOLVED'] = {
        'IF_REACHED_WATER': agate.Text(), # CY95 UNKNOWN
    }
    specified_types['MATERIAL_INV0LVED_CR'] = {
        'CAS_NUMBER': agate.Text(), # CY97 000000-00-0
        'UPPER_BOUNDS': agate.Text(), # CY05 UNKNOWN
    }
    specified_types['MOBILE_DETAILS'] = {
        'CARGO_CAPACITY': agate.Text(), # CY02 O instead of 0
    }
    specified_types['TRAINS_DETAIL'] = {
        'NON_COMPLIANCE_WITH_HAZMAT': agate.Text(),
    }
    specified_types['VESSELS_DETAIL'] = {
        'CARGO_ON_BOARD': agate.Text(), # CY06 SEWAGE
        'FUEL_CAPACITY': agate.Text(), # CY03 110,000 (that comma...)
        'FUEL_ON_BOARD': agate.Text(), # CY03 53,000 (that comma...)
    }

    for sheetname in sheetnames:
        print(sheetname)
        start = datetime.now()
        t = agate.Table.from_xlsx(filename,
                                  sheet=sheetname,
                                  column_types=specified_types.get(sheetname, {}))
        # we create a duplicate table with lowercase column names because
        # uppercase names require quoting in resulting SQL statements
        # and we lowercase the SQL table name for the same reason in the following
        # line
        t = t.rename(column_names = [name.lower() for name in t.column_names])
        t.to_sql(sqlurl,
                 sheetname.lower(),
                 constraints=False,
                 create=True,
                 create_if_not_exists=True,
                 # chunk_size=1,
        )
        delta = datetime.now() - start
        print("\t%d" % delta.seconds)

def extractor_command():
    """
    This command extracts data from National Response Center (NRC) incident archive
    spreadsheets into programmer friendly JSON documents.
    """

    parser = OptionParser(usage="usage: %%prog [options]\n%s" % extractor_command.__doc__)
    parser.add_option("-f", "--input-file", dest="input_file")
    parser.add_option("-i", "--input-directory", dest="input_directory",
                      help="Parse all files in the provided directory")

    (options, args) = parser.parse_args()

    if options.input_file is None and options.input_directory is None:
        print("You must supply an input file or input directory")
        parser.print_help()
        sys.exit(-1)

    file_paths = []

    if options.input_file:
        if not os.path.exists(options.input_file):
            print("The input file does not exist")
            sys.exit(-1)
        else:
            file_paths.append(os.path.abspath(options.input_file))

    if options.input_directory:
        input_dir = options.input_directory
        if not os.path.exists(input_dir):
            print("The input directory does not exist")
            sys.exit(-1)
        else:
            input_dir = os.path.abspath(input_dir)
            for name in os.listdir(input_dir):
                file_paths.append(os.path.sep.join((input_dir, name)))

    for file_path in file_paths:
        print(file_path)
        extract_xlsx_to_sql(file_path, env_db_string())

    # for file_path in file_paths:
    #     extractor = None
    #     for extract_class in extractors:
    #         if extract_class.mimetype == mimetypes.guess_type(file_path)[0]:
    #             extractor = extract_class(file_path)
                
    #     if extractor is None:
    #         print("No extractor available for %s" % file_path)
    #     else:
    #         extractor.extract_data(data_storage)
            
    print("Extracted data from NRC spreadsheets")
